from openpyxl import Workbook,load_workbook
import pymysql,re


import configparser,os
cf = configparser.ConfigParser()
cf.read("trans_conf.ini")

class sql_trans:
    def __init__(self) -> None:

        
        host1=cf.get("mysql", "host")
        port1=cf.get("mysql", "port")
        db_name1=cf.get("mysql", "db_name")
        user_name1=cf.get("mysql", "user_name")
        password1=cf.get("mysql", "password")
        self.file_dir=cf.get("mysql", "file_dir")
        self.file_output_file=cf.get("mysql", "file_output_file")
        self.file_list=os.listdir(self.file_dir)
        if not self.file_list:
            raise Exception('{}文件夹内无文件，请放入后重试'.format(self.file_dir))
        self.db = pymysql.connect(host=host1,port=int(port1),user=user_name1,password=password1,database=db_name1)
        self.cursor = self.db.cursor()

        self.region_dict={}
        region_sql="SELECT * FROM dict_region"
        self.cursor.execute(region_sql)
        results = self.cursor.fetchall()
        for r in results:
            region_code=r[2]
            region_name=r[1]
            self.region_dict[region_code]=region_name
    def gao(self):


        for ff in self.file_list:


            wb = load_workbook(self.file_dir+'/'+ff)
            ws=wb.active
            header1=list(ws.rows)[0]
            header1_list=[x.internal_value for x in header1][1:]
            header2=list(ws.rows)[1]
            header2_list=[x.internal_value for x in header2][1:]
            info_dict={}
            header_list=[x.internal_value for x in header1][1:]
            sql = "SELECT * FROM tjk_indicator_base \
                WHERE indicator_name = '%s' and year = '%s'" 
            year=re.findall('\d{4}',ff)
            if not year:
                raise Exception("没有找到年份")#异常被抛出，print函数无法执行
            year=year[0]
            for header in header_list:
                sql_row=sql % (header,year)
                self.cursor.execute(sql_row)
                results = self.cursor.fetchall()

                if results:
                    for r in results:
                        region_code=r[4]
                        region_name=self.region_dict[region_code]
                        if not region_name in info_dict:
                            info_dict[region_name]={}

                        indicator_value=r[8]
                        info_dict[region_name][header]=indicator_value
            score_list=[]
            header_dict={}
            for head in header_list:
                header_dict[head]=''

            wb_new=Workbook()
            ws_new=wb_new.active
            header1_list.insert(0,'key')
            header2_list.insert(0,'来源')
            ws_new.append(header1_list)
            ws_new.append(header2_list)
            for region in info_dict:
                header_dict2=header_dict
                indicator_info=info_dict[region]
                row=[region]
                for indicator in indicator_info:
                    score=indicator_info[indicator]
                    header_dict2[indicator]=score
                values=header_dict2.values()
                row+=list(values)
                ws_new.append(row)
            wb_new.save(self.file_output_file+'/output''_'+ff)


if __name__=='__main__':
    sql_t=sql_trans()
    sql_t.gao()